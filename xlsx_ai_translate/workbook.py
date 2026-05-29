from __future__ import annotations

import math
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Iterator, Sequence

from .ai import TranslationClient, TranslationError


@dataclass
class TranslationStats:
    sheets_seen: int = 0
    sheets_translated: int = 0
    excluded_sheets: list[str] = field(default_factory=list)
    missing_excluded_sheets: list[str] = field(default_factory=list)
    cells_scanned: int = 0
    formula_cells_skipped: int = 0
    blank_strings_skipped: int = 0
    string_cells_found: int = 0
    unique_strings: int = 0
    duplicate_strings_reused: int = 0


@dataclass(frozen=True)
class _CellText:
    cell: object
    text: str


def translate_workbook(
    input_path: str | Path,
    output_path: str | Path,
    *,
    client: TranslationClient,
    target_language: str,
    source_language: str = "auto",
    exclude_sheets: Sequence[str] = (),
    batch_size: int = 50,
    max_batch_chars: int = 12_000,
    concurrency: int = 8,
    requests_per_minute: int | None = 500,
    tokens_per_minute: int | None = 200_000,
    progress: Callable[[int, int], None] | None = None,
) -> TranslationStats:
    """Translate string cells in an XLSX workbook and save the translated copy."""
    from openpyxl import load_workbook

    input_file = Path(input_path)
    output_file = Path(output_path)
    _validate_paths(input_file, output_file)
    if batch_size < 1:
        raise ValueError("batch_size must be at least 1.")
    if max_batch_chars < 1:
        raise ValueError("max_batch_chars must be at least 1.")
    if concurrency < 1:
        raise ValueError("concurrency must be at least 1.")
    if requests_per_minute is not None and requests_per_minute < 1:
        raise ValueError("requests_per_minute must be at least 1.")
    if tokens_per_minute is not None and tokens_per_minute < 1:
        raise ValueError("tokens_per_minute must be at least 1.")

    workbook = load_workbook(input_file, data_only=False)
    excluded = set(exclude_sheets)
    stats = TranslationStats(
        sheets_seen=len(workbook.worksheets),
        missing_excluded_sheets=sorted(excluded - set(workbook.sheetnames)),
    )

    cells_to_translate: list[_CellText] = []
    unique_texts: list[str] = []
    translation_cache: dict[str, str | None] = {}

    for sheet in workbook.worksheets:
        if sheet.title in excluded:
            stats.excluded_sheets.append(sheet.title)
            continue

        stats.sheets_translated += 1
        for row in sheet.iter_rows():
            for cell in row:
                stats.cells_scanned += 1
                value = cell.value
                if not isinstance(value, str):
                    continue
                if cell.data_type == "f":
                    stats.formula_cells_skipped += 1
                    continue
                if value == "" or value.isspace():
                    stats.blank_strings_skipped += 1
                    continue

                stats.string_cells_found += 1
                cells_to_translate.append(_CellText(cell=cell, text=value))
                if value not in translation_cache:
                    translation_cache[value] = None
                    unique_texts.append(value)

    stats.unique_strings = len(unique_texts)
    stats.duplicate_strings_reused = stats.string_cells_found - stats.unique_strings

    batches = list(_batches(unique_texts, max_items=batch_size, max_chars=max_batch_chars))
    batch_results = _translate_batches(
        batches,
        client=client,
        source_language=source_language,
        target_language=target_language,
        concurrency=concurrency,
        requests_per_minute=requests_per_minute,
        tokens_per_minute=tokens_per_minute,
        progress=progress,
    )
    for batch, translations in zip(batches, batch_results, strict=True):
        for source, translated in zip(batch, translations, strict=True):
            translation_cache[source] = translated

    for item in cells_to_translate:
        translated = translation_cache[item.text]
        if translated is None:
            raise TranslationError(f"Missing translation for cell text: {item.text!r}")
        item.cell.value = translated

    workbook.save(output_file)
    return stats


def _validate_paths(input_file: Path, output_file: Path) -> None:
    if input_file.suffix.lower() != ".xlsx":
        raise ValueError("input_path must point to a .xlsx file.")
    if output_file.suffix.lower() != ".xlsx":
        raise ValueError("output_path must point to a .xlsx file.")
    if not input_file.exists():
        raise FileNotFoundError(input_file)
    if not output_file.parent.exists():
        raise FileNotFoundError(output_file.parent)


def _batches(
    items: Sequence[str],
    *,
    max_items: int,
    max_chars: int,
) -> Iterator[list[str]]:
    batch: list[str] = []
    batch_chars = 0

    for item in items:
        item_chars = len(item)
        if batch and (len(batch) >= max_items or batch_chars + item_chars > max_chars):
            yield batch
            batch = []
            batch_chars = 0

        batch.append(item)
        batch_chars += item_chars

        if item_chars >= max_chars:
            yield batch
            batch = []
            batch_chars = 0

    if batch:
        yield batch


def _translate_batches(
    batches: Sequence[Sequence[str]],
    *,
    client: TranslationClient,
    source_language: str,
    target_language: str,
    concurrency: int,
    requests_per_minute: int | None,
    tokens_per_minute: int | None,
    progress: Callable[[int, int], None] | None,
) -> list[list[str]]:
    if not batches:
        return []

    limiter = _RateLimiter(
        requests_per_minute=requests_per_minute,
        tokens_per_minute=tokens_per_minute,
    )
    batch_count = len(batches)
    results: list[list[str] | None] = [None] * batch_count

    def translate_batch(batch_index: int, batch: Sequence[str]) -> tuple[int, list[str]]:
        limiter.acquire(_estimate_tokens_for_batch(batch))
        if progress is not None:
            progress(batch_index + 1, batch_count)
        translations = client.translate_batch(
            batch,
            source_language=source_language,
            target_language=target_language,
        )
        if len(translations) != len(batch):
            raise TranslationError(
                f"Translator returned {len(translations)} translations for "
                f"{len(batch)} input strings in batch {batch_index + 1}."
            )
        return batch_index, translations

    max_workers = min(concurrency, batch_count)
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [
            executor.submit(translate_batch, batch_index, batch)
            for batch_index, batch in enumerate(batches)
        ]
        for future in as_completed(futures):
            batch_index, translations = future.result()
            results[batch_index] = translations

    missing = [index + 1 for index, result in enumerate(results) if result is None]
    if missing:
        raise TranslationError(f"Missing translations for batches: {missing}")
    return [result for result in results if result is not None]


class _RateLimiter:
    def __init__(
        self,
        *,
        requests_per_minute: int | None,
        tokens_per_minute: int | None,
    ) -> None:
        self.requests_per_minute = requests_per_minute
        self.tokens_per_minute = tokens_per_minute
        self._request_capacity = float(requests_per_minute or 0)
        self._token_capacity = float(tokens_per_minute or 0)
        self._updated_at = time.monotonic()
        self._lock = threading.Lock()

    def acquire(self, estimated_tokens: int) -> None:
        while True:
            with self._lock:
                self._refill()
                request_ready = (
                    self.requests_per_minute is None or self._request_capacity >= 1
                )
                token_cost = self._token_cost(estimated_tokens)
                token_ready = self.tokens_per_minute is None or self._token_capacity >= token_cost

                if request_ready and token_ready:
                    if self.requests_per_minute is not None:
                        self._request_capacity -= 1
                    if self.tokens_per_minute is not None:
                        self._token_capacity -= token_cost
                    return

                wait_seconds = self._wait_seconds(token_cost)

            time.sleep(max(0.05, min(wait_seconds, 1.0)))

    def _refill(self) -> None:
        now = time.monotonic()
        elapsed = now - self._updated_at
        self._updated_at = now

        if self.requests_per_minute is not None:
            self._request_capacity = min(
                float(self.requests_per_minute),
                self._request_capacity + self.requests_per_minute * elapsed / 60,
            )
        if self.tokens_per_minute is not None:
            self._token_capacity = min(
                float(self.tokens_per_minute),
                self._token_capacity + self.tokens_per_minute * elapsed / 60,
            )

    def _token_cost(self, estimated_tokens: int) -> float:
        if self.tokens_per_minute is None:
            return 0
        return float(min(max(estimated_tokens, 1), self.tokens_per_minute))

    def _wait_seconds(self, token_cost: float) -> float:
        waits: list[float] = []
        if self.requests_per_minute is not None and self._request_capacity < 1:
            waits.append((1 - self._request_capacity) * 60 / self.requests_per_minute)
        if self.tokens_per_minute is not None and self._token_capacity < token_cost:
            waits.append((token_cost - self._token_capacity) * 60 / self.tokens_per_minute)
        return max(waits) if waits else 0.05


def _estimate_tokens_for_batch(batch: Sequence[str]) -> int:
    source_chars = sum(len(item) for item in batch)
    json_overhead_chars = 500 + 12 * len(batch)
    estimated_input_tokens = math.ceil((source_chars + json_overhead_chars) / 4)
    return max(1, estimated_input_tokens * 2)


def stderr_progress(batch_index: int, batch_count: int) -> None:
    print(f"Translating batch {batch_index}/{batch_count}...", file=sys.stderr)
