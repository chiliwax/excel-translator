from __future__ import annotations

import threading
import time

from openpyxl import Workbook, load_workbook

from xlsx_ai_translate.workbook import _batches, _estimate_tokens_for_batch, translate_workbook


class FakeTranslationClient:
    def __init__(self) -> None:
        self.calls: list[list[str]] = []

    def translate_batch(
        self,
        texts: list[str],
        *,
        source_language: str,
        target_language: str,
    ) -> list[str]:
        self.calls.append(list(texts))
        return [f"{text} -> {target_language}" for text in texts]


def test_translates_unique_strings_once_and_preserves_formulas(tmp_path):
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Main"
    sheet["A1"] = "Hello"
    sheet["A2"] = "Hello"
    sheet["A3"] = "hello"
    sheet["B1"] = "=SUM(1,2)"
    sheet["B2"] = 123
    sheet["C1"] = " "
    workbook.save(input_path)

    client = FakeTranslationClient()
    stats = translate_workbook(
        input_path,
        output_path,
        client=client,
        target_language="fr",
        batch_size=1,
        concurrency=1,
        translate_sheet_names=False,
    )

    translated = load_workbook(output_path, data_only=False)
    translated_sheet = translated["Main"]

    assert translated_sheet["A1"].value == "Hello -> fr"
    assert translated_sheet["A2"].value == "Hello -> fr"
    assert translated_sheet["A3"].value == "hello -> fr"
    assert translated_sheet["B1"].value == "=SUM(1,2)"
    assert translated_sheet["B2"].value == 123
    assert translated_sheet["C1"].value == " "
    assert client.calls == [["Hello"], ["hello"]]
    assert stats.string_cells_found == 3
    assert stats.unique_strings == 2
    assert stats.duplicate_strings_reused == 1
    assert stats.formula_cells_skipped == 1
    assert stats.blank_strings_skipped == 1


def test_excludes_requested_sheets_and_reports_missing_names(tmp_path):
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"

    workbook = Workbook()
    keep = workbook.active
    keep.title = "Keep"
    skip = workbook.create_sheet("Skip")
    keep["A1"] = "Hello"
    skip["A1"] = "Hello"
    workbook.save(input_path)

    client = FakeTranslationClient()
    stats = translate_workbook(
        input_path,
        output_path,
        client=client,
        target_language="es",
        exclude_sheets=["Skip", "Missing"],
        translate_sheet_names=False,
    )

    translated = load_workbook(output_path, data_only=False)

    assert translated["Keep"]["A1"].value == "Hello -> es"
    assert translated["Skip"]["A1"].value == "Hello"
    assert client.calls == [["Hello"]]
    assert stats.sheets_seen == 2
    assert stats.sheets_translated == 1
    assert stats.excluded_sheets == ["Skip"]
    assert stats.missing_excluded_sheets == ["Missing"]


def test_translates_all_sheets_by_default(tmp_path):
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"

    workbook = Workbook()
    first = workbook.active
    first.title = "First"
    second = workbook.create_sheet("Second")
    first["A1"] = "One"
    second["A1"] = "Two"
    workbook.save(input_path)

    client = FakeTranslationClient()
    translate_workbook(
        input_path,
        output_path,
        client=client,
        target_language="de",
        translate_sheet_names=False,
    )

    translated = load_workbook(output_path, data_only=False)

    assert translated["First"]["A1"].value == "One -> de"
    assert translated["Second"]["A1"].value == "Two -> de"
    assert client.calls == [["One", "Two"]]


def test_batches_by_count_and_character_budget():
    assert list(_batches(["aa", "bb", "cccc", "d"], max_items=3, max_chars=5)) == [
        ["aa", "bb"],
        ["cccc", "d"],
    ]


def test_batches_single_large_item_alone():
    assert list(_batches(["abcdef", "g"], max_items=10, max_chars=5)) == [
        ["abcdef"],
        ["g"],
    ]


def test_parallel_translation_respects_concurrency_and_maps_results(tmp_path):
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "One"
    sheet["A2"] = "Two"
    sheet["A3"] = "Three"
    sheet["A4"] = "Four"
    workbook.save(input_path)

    client = TrackingTranslationClient()
    translate_workbook(
        input_path,
        output_path,
        client=client,
        target_language="en",
        batch_size=1,
        concurrency=2,
        requests_per_minute=500,
        tokens_per_minute=200_000,
        translate_sheet_names=False,
    )

    translated = load_workbook(output_path, data_only=False)
    sheet = translated.active

    assert [sheet[f"A{row}"].value for row in range(1, 5)] == [
        "One -> en",
        "Two -> en",
        "Three -> en",
        "Four -> en",
    ]
    assert client.max_active == 2


def test_estimate_tokens_for_batch_accounts_for_output():
    assert _estimate_tokens_for_batch(["a" * 400]) >= 200


def test_translates_sheet_titles_with_same_cache_and_updates_formula_references(tmp_path):
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"

    workbook = Workbook()
    data = workbook.active
    data.title = "Data"
    summary = workbook.create_sheet("Summary")
    data["A1"] = "Data"
    summary["A1"] = "=Data!A1"
    workbook.save(input_path)

    client = MappingTranslationClient({"Data": "Donnees", "Summary": "Resume"})
    stats = translate_workbook(input_path, output_path, client=client, target_language="fr")

    translated = load_workbook(output_path, data_only=False)

    assert translated.sheetnames == ["Donnees", "Resume"]
    assert translated["Donnees"]["A1"].value == "Donnees"
    assert translated["Resume"]["A1"].value == "='Donnees'!A1"
    assert client.calls == [["Data", "Summary"]]
    assert stats.sheet_titles_found == 2
    assert stats.sheet_titles_translated == 2
    assert stats.unique_strings == 2


def test_excluded_sheet_title_is_not_translated(tmp_path):
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"

    workbook = Workbook()
    keep = workbook.active
    keep.title = "Keep"
    skip = workbook.create_sheet("Skip")
    keep["A1"] = "Hello"
    skip["A1"] = "Hello"
    workbook.save(input_path)

    client = MappingTranslationClient({"Keep": "Garder", "Hello": "Bonjour"})
    stats = translate_workbook(
        input_path,
        output_path,
        client=client,
        target_language="fr",
        exclude_sheets=["Skip"],
    )

    translated = load_workbook(output_path, data_only=False)

    assert translated.sheetnames == ["Garder", "Skip"]
    assert translated["Garder"]["A1"].value == "Bonjour"
    assert translated["Skip"]["A1"].value == "Hello"
    assert stats.sheet_titles_found == 1
    assert stats.sheet_titles_translated == 1


def test_translated_sheet_titles_are_valid_unique_excel_names(tmp_path):
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"

    workbook = Workbook()
    first = workbook.active
    first.title = "First"
    second = workbook.create_sheet("Second")
    first["A1"] = "A"
    second["A1"] = "B"
    workbook.save(input_path)

    long_invalid_name = "Very/Long:Translated*Worksheet?Name With Extra Words"
    client = MappingTranslationClient(
        {
            "First": long_invalid_name,
            "Second": long_invalid_name,
            "A": "AA",
            "B": "BB",
        }
    )
    translate_workbook(input_path, output_path, client=client, target_language="en")

    translated = load_workbook(output_path, data_only=False)

    assert len(translated.sheetnames) == len(set(translated.sheetnames))
    assert all(len(title) <= 31 for title in translated.sheetnames)
    assert all(not set(title) & set('\\/*?:[]') for title in translated.sheetnames)


class TrackingTranslationClient:
    def __init__(self) -> None:
        self.active = 0
        self.max_active = 0
        self.lock = threading.Lock()

    def translate_batch(
        self,
        texts: list[str],
        *,
        source_language: str,
        target_language: str,
    ) -> list[str]:
        with self.lock:
            self.active += 1
            self.max_active = max(self.max_active, self.active)
        try:
            time.sleep(0.05)
            return [f"{text} -> {target_language}" for text in texts]
        finally:
            with self.lock:
                self.active -= 1


class MappingTranslationClient:
    def __init__(self, translations: dict[str, str]) -> None:
        self.translations = translations
        self.calls: list[list[str]] = []

    def translate_batch(
        self,
        texts: list[str],
        *,
        source_language: str,
        target_language: str,
    ) -> list[str]:
        self.calls.append(list(texts))
        return [self.translations.get(text, f"{text} -> {target_language}") for text in texts]
