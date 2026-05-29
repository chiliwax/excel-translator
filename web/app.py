from __future__ import annotations

import json
import os
import re
import shutil
import threading
import uuid
from contextlib import asynccontextmanager
from concurrent.futures import ThreadPoolExecutor
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Annotated

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from xlsx_ai_translate.ai import LiteLLMTranslationClient
from xlsx_ai_translate.cli import _load_env_file
from xlsx_ai_translate.workbook import TranslationStats, translate_workbook


APP_ROOT = Path(__file__).resolve().parent
PROJECT_ROOT = APP_ROOT.parent
DATA_DIR = Path(os.environ.get("XLSX_TRANSLATOR_DATA_DIR", PROJECT_ROOT / ".data")).resolve()
JOBS_DIR = DATA_DIR / "jobs"
UPLOADS_DIR = DATA_DIR / "uploads"
MAX_UPLOAD_BYTES = int(os.environ.get("XLSX_TRANSLATOR_MAX_UPLOAD_MB", "100")) * 1024 * 1024
JOB_EXECUTOR = ThreadPoolExecutor(max_workers=int(os.environ.get("XLSX_TRANSLATOR_WEB_WORKERS", "2")))
_SAFE_ID = re.compile(r"^[0-9a-f]{32}$")

@asynccontextmanager
async def lifespan(_: FastAPI):
    env_file = Path(os.environ.get("ENV_FILE", PROJECT_ROOT / ".env"))
    _load_env_file(env_file)
    JOBS_DIR.mkdir(parents=True, exist_ok=True)
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    yield


app = FastAPI(title="XLSX AI Translate", lifespan=lifespan)
templates = Jinja2Templates(directory=str(APP_ROOT / "templates"))
app.mount("/static", StaticFiles(directory=str(APP_ROOT / "static")), name="static")

_jobs: dict[str, "JobState"] = {}
_jobs_lock = threading.Lock()


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


@dataclass
class JobState:
    id: str
    status: str
    input_filename: str
    source_language: str
    target_language: str
    model: str
    exclude_sheets: list[str]
    translate_sheet_names: bool
    batch_size: int
    max_batch_chars: int
    concurrency: int
    rpm: int
    tpm: int
    request_timeout: float
    created_at: str = field(default_factory=_now_iso)
    updated_at: str = field(default_factory=_now_iso)
    progress_current: int = 0
    progress_total: int = 0
    message: str = "Queued"
    error: str | None = None
    output_filename: str | None = None
    stats: dict[str, object] | None = None

    def public_dict(self) -> dict[str, object]:
        data = asdict(self)
        data["download_url"] = f"/jobs/{self.id}/download" if self.status == "done" else None
        data["progress_percent"] = self.progress_percent
        return data

    @property
    def progress_percent(self) -> int:
        if self.status == "done":
            return 100
        if self.progress_total <= 0:
            return 0
        return min(99, round(self.progress_current * 100 / self.progress_total))


@app.get("/", response_class=HTMLResponse)
def index(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        request=request,
        name="index.html",
        context={
            "request": request,
            "default_model": os.environ.get("XLSX_TRANSLATOR_MODEL", "openai/gpt-4o-mini"),
            "max_upload_mb": MAX_UPLOAD_BYTES // 1024 // 1024,
        },
    )


@app.post("/inspect")
def inspect_workbook(file: Annotated[UploadFile, File()]) -> JSONResponse:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Upload must be a .xlsx file.")

    upload_id = uuid.uuid4().hex
    upload_dir = UPLOADS_DIR / upload_id
    upload_dir.mkdir(parents=True)
    input_path = upload_dir / "input.xlsx"
    _save_upload(file, input_path)

    try:
        sheets = _read_sheet_names(input_path)
    except Exception as exc:
        shutil.rmtree(upload_dir, ignore_errors=True)
        raise HTTPException(status_code=400, detail=f"Could not inspect workbook: {exc}") from exc

    metadata = {
        "upload_id": upload_id,
        "filename": file.filename,
        "sheets": sheets,
        "created_at": _now_iso(),
    }
    (upload_dir / "upload.json").write_text(json.dumps(metadata, indent=2), encoding="utf-8")
    return JSONResponse(metadata)


@app.post("/jobs")
def create_job(
    target_language: Annotated[str, Form()],
    file: Annotated[UploadFile | None, File()] = None,
    upload_id: Annotated[str, Form()] = "",
    source_language: Annotated[str, Form()] = "auto",
    model: Annotated[str, Form()] = "",
    exclude_sheets: Annotated[str, Form()] = "",
    exclude_sheet_names: Annotated[list[str] | None, Form()] = None,
    translate_sheet_names: Annotated[bool, Form()] = False,
    batch_size: Annotated[int, Form()] = 50,
    max_batch_chars: Annotated[int, Form()] = 12_000,
    concurrency: Annotated[int, Form()] = 8,
    rpm: Annotated[int, Form()] = 500,
    tpm: Annotated[int, Form()] = 200_000,
    request_timeout: Annotated[float, Form()] = 120.0,
) -> RedirectResponse:
    if not target_language.strip():
        raise HTTPException(status_code=400, detail="Target language is required.")

    _validate_positive("batch_size", batch_size)
    _validate_positive("max_batch_chars", max_batch_chars)
    _validate_positive("concurrency", concurrency)
    _validate_positive("rpm", rpm)
    _validate_positive("tpm", tpm)
    if request_timeout <= 0:
        raise HTTPException(status_code=400, detail="request_timeout must be positive.")

    job_id = uuid.uuid4().hex
    job_dir = JOBS_DIR / job_id
    job_dir.mkdir(parents=True)

    input_path = job_dir / "input.xlsx"
    staged_input_path, input_filename = _resolve_job_input(upload_id, file)
    output_filename = f"{Path(input_filename).stem}.translated.xlsx"
    output_path = job_dir / output_filename
    shutil.copyfile(staged_input_path, input_path)

    parsed_excluded_sheets = _parse_exclude_sheets(exclude_sheets)
    for sheet_name in exclude_sheet_names or []:
        if sheet_name and sheet_name not in parsed_excluded_sheets:
            parsed_excluded_sheets.append(sheet_name)

    job = JobState(
        id=job_id,
        status="queued",
        input_filename=input_filename,
        source_language=source_language.strip() or "auto",
        target_language=target_language.strip(),
        model=model.strip() or os.environ.get("XLSX_TRANSLATOR_MODEL", "openai/gpt-4o-mini"),
        exclude_sheets=parsed_excluded_sheets,
        translate_sheet_names=translate_sheet_names,
        batch_size=batch_size,
        max_batch_chars=max_batch_chars,
        concurrency=concurrency,
        rpm=rpm,
        tpm=tpm,
        request_timeout=request_timeout,
        output_filename=output_filename,
    )
    _set_job(job)
    _write_job_file(job)
    JOB_EXECUTOR.submit(_run_job, job_id, input_path, output_path)

    return RedirectResponse(url=f"/jobs/{job_id}", status_code=303)


@app.get("/jobs/{job_id}", response_class=HTMLResponse)
def job_page(request: Request, job_id: str) -> HTMLResponse:
    job = _get_job(job_id)
    return templates.TemplateResponse(
        request=request,
        name="job.html",
        context={"request": request, "job": job.public_dict()},
    )


@app.get("/jobs/{job_id}/status")
def job_status(job_id: str) -> JSONResponse:
    job = _get_job(job_id)
    return JSONResponse(job.public_dict())


@app.get("/jobs/{job_id}/download")
def download(job_id: str) -> FileResponse:
    job = _get_job(job_id)
    if job.status != "done" or not job.output_filename:
        raise HTTPException(status_code=404, detail="Translated workbook is not ready.")

    output_path = JOBS_DIR / job_id / job.output_filename
    if not output_path.exists():
        raise HTTPException(status_code=404, detail="Translated workbook was not found.")

    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=job.output_filename,
    )


def _run_job(job_id: str, input_path: Path, output_path: Path) -> None:
    job = _get_job(job_id)
    _update_job(job_id, status="running", message="Starting translation")

    def progress(current: int, total: int) -> None:
        _update_job(
            job_id,
            progress_current=current,
            progress_total=total,
            message=f"Translated batch {current}/{total}",
        )

    try:
        client = LiteLLMTranslationClient(
            model=job.model,
            temperature=0.0,
            timeout_seconds=job.request_timeout,
        )
        stats = translate_workbook(
            input_path,
            output_path,
            client=client,
            source_language=job.source_language,
            target_language=job.target_language,
            exclude_sheets=job.exclude_sheets,
            batch_size=job.batch_size,
            max_batch_chars=job.max_batch_chars,
            concurrency=job.concurrency,
            requests_per_minute=job.rpm,
            tokens_per_minute=job.tpm,
            translate_sheet_names=job.translate_sheet_names,
            progress=progress,
        )
        _update_job(
            job_id,
            status="done",
            progress_current=max(job.progress_current, job.progress_total),
            message="Translation complete",
            stats=_stats_dict(stats),
        )
    except Exception as exc:
        _update_job(job_id, status="error", error=str(exc), message="Translation failed")


def _save_upload(file: UploadFile, destination: Path) -> None:
    bytes_written = 0
    with destination.open("wb") as output:
        while True:
            chunk = file.file.read(1024 * 1024)
            if not chunk:
                break
            bytes_written += len(chunk)
            if bytes_written > MAX_UPLOAD_BYTES:
                raise HTTPException(status_code=413, detail="Uploaded file is too large.")
            output.write(chunk)


def _parse_exclude_sheets(value: str) -> list[str]:
    return [item.strip() for item in value.replace("\n", ",").split(",") if item.strip()]


def _read_sheet_names(path: Path) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=False)
    return list(workbook.sheetnames)


def _resolve_job_input(upload_id: str, file: UploadFile | None) -> tuple[Path, str]:
    if upload_id:
        safe_upload_id = _safe_data_id(upload_id)
        upload_dir = UPLOADS_DIR / safe_upload_id
        input_path = upload_dir / "input.xlsx"
        metadata_path = upload_dir / "upload.json"
        if not input_path.exists() or not metadata_path.exists():
            raise HTTPException(status_code=404, detail="Inspected upload was not found.")
        metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        return input_path, metadata.get("filename") or "workbook.xlsx"

    if file is None or not file.filename:
        raise HTTPException(status_code=400, detail="Inspect and select a workbook before starting translation.")
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Upload must be a .xlsx file.")

    temp_upload_id = uuid.uuid4().hex
    upload_dir = UPLOADS_DIR / temp_upload_id
    upload_dir.mkdir(parents=True)
    input_path = upload_dir / "input.xlsx"
    _save_upload(file, input_path)
    return input_path, file.filename


def _safe_data_id(value: str) -> str:
    if not _SAFE_ID.fullmatch(value):
        raise HTTPException(status_code=400, detail="Invalid upload id.")
    return value


def _validate_positive(name: str, value: int) -> None:
    if value < 1:
        raise HTTPException(status_code=400, detail=f"{name} must be at least 1.")


def _set_job(job: JobState) -> None:
    with _jobs_lock:
        _jobs[job.id] = job


def _get_job(job_id: str) -> JobState:
    with _jobs_lock:
        job = _jobs.get(job_id)
    if job is None:
        job = _load_job_file(job_id)
        if job is None:
            raise HTTPException(status_code=404, detail="Job not found.")
        _set_job(job)
    return job


def _update_job(job_id: str, **changes: object) -> None:
    with _jobs_lock:
        job = _jobs[job_id]
        for key, value in changes.items():
            setattr(job, key, value)
        job.updated_at = _now_iso()
    _write_job_file(job)


def _write_job_file(job: JobState) -> None:
    job_dir = JOBS_DIR / job.id
    job_dir.mkdir(parents=True, exist_ok=True)
    temp_path = job_dir / "job.json.tmp"
    final_path = job_dir / "job.json"
    temp_path.write_text(json.dumps(job.public_dict(), indent=2), encoding="utf-8")
    temp_path.replace(final_path)


def _load_job_file(job_id: str) -> JobState | None:
    job_file = JOBS_DIR / job_id / "job.json"
    if not job_file.exists():
        return None
    data = json.loads(job_file.read_text(encoding="utf-8"))
    allowed_keys = {field.name for field in JobState.__dataclass_fields__.values()}
    job = JobState(**{key: value for key, value in data.items() if key in allowed_keys})
    if job.status in {"queued", "running"}:
        job.status = "error"
        job.message = "Job was interrupted by server restart."
        job.error = job.message
        _write_job_file(job)
    return job


def _stats_dict(stats: TranslationStats) -> dict[str, object]:
    return asdict(stats)


def reset_for_tests(data_dir: Path) -> None:
    global DATA_DIR, JOBS_DIR, UPLOADS_DIR
    with _jobs_lock:
        _jobs.clear()
    DATA_DIR = data_dir.resolve()
    JOBS_DIR = DATA_DIR / "jobs"
    UPLOADS_DIR = DATA_DIR / "uploads"
    if DATA_DIR.exists():
        shutil.rmtree(DATA_DIR)
    JOBS_DIR.mkdir(parents=True, exist_ok=True)
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
